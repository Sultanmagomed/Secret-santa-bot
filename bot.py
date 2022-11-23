#!/usr/bin/python
# -*- coding: utf8 -*-
from pydoc import cli
import discord
from discord.ext import commands
from openpyxl import load_workbook
import shutil
import os
from datetime import datetime
from typing import Literal


PREFIX = '!' #префикс для всяких комманд, которые не используются
fn = str(os.path.dirname(__file__))+'/List_of_santas.xlsx' #xl файл для записи участников и реквестов
lf = str(os.path.dirname(__file__))+'/log.txt'#txt файл логов
lfbak = str(os.path.dirname(__file__)) + '/log.bak.txt'#txt файл предыдущих логов
#bot_master = 'bespilotnik#7796' #кто тут истиный данжнмастер, хотя нужен именно id так что эта строчка для наглядности
cdtime = datetime.now() #текущая датавремя
try:
    os.remove(lfbak) #удаление старых логов
except:
    print('Файл старых логов не существовал и не удалён')
try:
    shutil.copy(lf, lfbak) #сохранение логов предыдущего запуска
except:
    print('файл логов не сущестовал, создан новый файл')
with open(lf,'w') as log:#файл логирования консоли
    log.write(str(cdtime)+' создан логфайл\n')
bot = commands.Bot(command_prefix = PREFIX, help_command=None, intents=discord.Intents.all())
bot_master_id = int(401465528004116481)#кто тут истиный данжнмастер в скобках id пользователя, управляющего ботом
bot.remove_command('help')

#словари
hello_words = ['sup', 'hi', 'hellow', 'привет', 'сап', 'драсте', 'приветствую']
answer_words = ['как дел', 'как дела', 'который час', 'какой сейчас год','как жить эту ебаную жизнь','в чем смысл жизни']
paricipant_word = ['участвовать', 'хочу участвовать', 'быть сантой', 'хочу быть сантой', 'запишите', 'запиши', 'запишите меня', 'запиши меня', 'want to participate','participate']
adding_words = ['добавить', 'дополнить', 'добавь', 'дополни', 'дополнение', 'ps', 'add', 'add request', 'ещё', 'ещё кое-что', 'ещё кое что', 'ещё коечто',]
yes_words = ['да', 'lf', 'ja', 'yes', 'es', 'угу']
no_words = ['нет', 'не', 'ytn', 'no', 'нит']
help_words = ['help','помогай','помощь','помогит','помогите','объясни','комманды','что делать','как участвовать','харп','херп','хелп','harp','herp','объясните']
rules_words = ['правила', 'rulz', 'rulez', 'rules', 'как играть', 'как быть сантой','как быть секретным сантой','как жить эту жизнь','рулз','рул']
recive_words = ['прими результат', 'результат','отправить результат','рисуночек','рисунок','прими рисунок','принемай']

#сообщение о готовности и указание статуса бота
@bot.event
async def on_ready():
    print()
    print("Успешно залогинились как {0.user} ".format(bot)) 
    await bot.change_presence( status=discord.Status.online, activity=discord.Game('секретного санту'))
    with open(lf,'a') as log:
        log.write(str(cdtime)+" успешно залогинились как {0.user} ".format(bot)+'\n')
    #print(str(os.path.dirname(os.path.dirname(__file__))))
    #print(str(os.path.dirname(__file__)))

#функция записи реквеста в файл
async def add_req(message, req):
    requestor = message.author.name + '#' + message.author.discriminator
    requestor_id = str(message.author.id)
    try:
        wb = load_workbook(fn)
        ws = wb['data']
        ws.append([requestor_id, requestor, req.content, str(await checknget_url(req))])
        wb.save(fn)
        wb.close
        print('успешная запиь реквеста в файл')
        with open(lf,'a') as log:
            log.write(str(cdtime)+" успешная запиь реквеста "+requestor+" в файл "+fn+'\n')
        await send_report(req, ' записался в участники')
    except:
        print('ошибка записи')
        with open(lf,'a') as log:
            log.write(str(cdtime)+" ошибка записи реквеста "+requestor+" в файл "+fn+'\n')
        await send_report(req, ' НЕ записался в участники, ошибка записи в файл')            
#функция дополнения реквеста
async def expand_req(req, str_number):
    requestor = req.author.name + '#' + req.author.discriminator
    try:
        wb = load_workbook(fn)
        ws = wb['data']
        if req.attachments and str(req.content) != '':                                                                                   #дополнение текстом при наличии картинки
            ws['C'+ str(str_number)] = str(wb.active.cell(row=str_number, column=3).value) + '. Дополнение: ' + req.content
        elif not req.attachments:                                                                                                        #дополнение текстом без картинки
            ws['C'+ str(str_number)] = str(wb.active.cell(row=str_number, column=3).value) + '. Дополнение: ' + req.content
        if str(wb.active.cell(row=str_number, column=4).value) == None or str(wb.active.cell(row=str_number, column=4).value) == 'None': #дополнение картинкой если не было картинок
            ws['D'+ str(str_number)] =str(await checknget_url(req))
        else:                                                                                                                            #дополнение картинкой если картинки были
            ws['D'+ str(str_number)] =str(wb.active.cell(row=str_number, column=4).value) + str(await checknget_url(req))
        wb.save(fn)
        wb.close
        print('успешная запись дополнения в файл')
        with open(lf,'a') as log:
            log.write(str(cdtime)+" успешная запиь дополнения "+requestor+" в файл "+fn+'\n')
        await send_report(req, ' дополнил реквест')
    except:
        print('ошибка записи дополнения в файл')
        with open(lf,'a') as log:
            log.write(str(cdtime)+" ошибка записи дополнения "+requestor+" в файл "+fn+'\n')
        await send_report(req, ' НЕ дополнил реквест, ошибка записи в файл')
#функция проверки участника на повторное участие/поиска номера участника как просителя и как рисователя
async def check_participy(message, status):
    if str(status) == 'requestor':
        col = 2
    elif str(status) == 'exequtor':
        col = 6
    requestor = message.author.name + '#' + message.author.discriminator
    print ('поиск '+requestor+' в столбце исполнителей')
    wb = load_workbook(fn)
    for i in range(1,500):
        value=wb.active.cell(row=i, column=col).value
        if value == requestor:
            print (requestor + 'исполнитель, строка №' + str(i))
            return i
    print (requestor + 'не участвовал ещё')
    wb.save(fn)
    wb.close
    return False
#служебная функция подсчёта количества участников
async def number_of_participant():
    print ('запрос количества учасиников')
    wb = load_workbook(fn)
    for i in range(1,500):
        value=wb.active.cell(row=i, column=2).value
        if str(value) == 'None':
            print ('участвует уже ' + str(i-2) + ' котов')
            participants = i-2
            wb.save(fn)
            wb.close
            return participants
#служебная функция рандомизации исполнителей (временно смещает участников на 1, что впрочем более чем случайно)
async def mixing_participant():
    wb = load_workbook(fn)
    boofer_name = wb.active.cell(row=2, column=2).value
    boofer_id = wb.active.cell(row=2, column=1).value
    for i in range(3,await number_of_participant()+3):
        wb.active.cell(row=i-1, column=6).value = wb.active.cell(row=i, column=2).value
        wb.active.cell(row=i-1, column=5).value = wb.active.cell(row=i, column=1).value
    wb.active.cell(row=i-1, column=6).value = boofer_name
    wb.active.cell(row=i-1, column=5).value = boofer_id
    wb.save(fn)
    wb.close
    return
#функция приёма выполненных реквестов (результатов)
async def recive_result(message, str_number):
    print ('получение результата от исполнителя')
    executor = message.author.name + '#' + message.author.discriminator
    try:
        wb = load_workbook(fn)
        ws = wb['data']
        if message.attachments and str(wb.active.cell(row=str_number, column=7).value) == 'None':
            ws['G'+ str(str_number)] = str(await checknget_url(message))
        elif message.attachments:
            ws['G'+ str(str_number)] = str(wb.active.cell(row=str_number, column=7).value) + str(await checknget_url(message))
        if str(message.content) != '' and str(wb.active.cell(row=str_number, column=8).value) == 'None':
            ws['H'+ str(str_number)] = message.content
        elif str(message.content) != '':
            ws['H'+ str(str_number)] = str(wb.active.cell(row=str_number, column=8).value) + '. Дополнение:' + message.content
        if str(wb.active.cell(row=str_number, column=8).value) == 'None':
            ws['H'+ str(str_number)] ='С новым годом!'
        wb.save(fn)
        wb.close
        print('успешная запись результата в файл')
        with open(lf,'a') as log:
            log.write(str(cdtime)+" успешная запиь результата от "+executor+" в файл "+fn+'\n')
        await send_report(message, ' отправил результат')
    except:
        print('ошибка записи результата в файл')
        with open(lf,'a') as log:
            log.write(str(cdtime)+" ошибка записи результата "+executor+" в файл "+fn+'\n')
        await send_report(message, ' НЕ отправил результат, ошибка записи в файл')
#служебная функция отправки результатов желателям
async def send_result():
    print ('запрос на отправку результатов заказчикам')
    try:
        wb = load_workbook(fn)
        for i in range(2,await number_of_participant()+2):
            requestor_id=int(wb.active.cell(row=i, column=1).value)
            result=wb.active.cell(row=i, column=7).value
            text_geeting=wb.active.cell(row=i, column=8).value
            await bot.get_user(requestor_id).send('Привет! Принимай поздравления и свой рисуночек')
            try:
                await bot.get_user(requestor_id).send(result+' '+text_geeting)
            except:
                try:
                    await bot.get_user(requestor_id).send(result)
                except:
                    await bot.get_user(requestor_id).send(text_geeting)
        wb.save(fn)
        wb.close
        with open(lf,'a') as log:
            log.write(str(cdtime)+" результаты отправленны желателям"+'\n')
    except:
        await bot.get_user(bot_master_id).send('не удалось отправить результаты')
        with open(lf,'a') as log:
            log.write(str(cdtime)+" не удалось отправить результаты желателям"+'\n')
    return
#функция получения url вложенных картинок
async def checknget_url(message):
    if  not message.guild and message.attachments:
        files = []
        for attachment in message.attachments:
            try:
                if attachment.content_type.startswith("image/"):
                    files.append(attachment.url)
            except:
                continue
        if files:
            return files
        else:
            return ''    
    else:
        return ''
#функция отправки сообщения владельцу при событии
async def send_report(message, report):
    await bot.get_user(bot_master_id).send(f'{message.author}'+report)        
#служебная функция отправления реквестов исполнителям
async def send_request():
    print ('запрос на отправку реквестов исполнителям')
    try:
        wb = load_workbook(fn)
        for i in range(2,await number_of_participant()+2):
            exequtor_id=int(wb.active.cell(row=i, column=5).value)
            request=wb.active.cell(row=i, column=3).value
            references=wb.active.cell(row=i, column=4).value
            await bot.get_user(exequtor_id).send('Привет! Принимай реквест. Как закончишь, скажи мне "прими результат" или просто "результат" и приложи картинку в сообщение и, если хочешь, можно какой-то текст, который получит поздравляемый. Помни, что срок до 30 декабря 2022 года. Удачи! \n Собственно реквест:')
            try:
                await bot.get_user(exequtor_id).send(request+' '+references)
            except:
                await bot.get_user(exequtor_id).send(request)
        wb.save(fn)
        wb.close
        with open(lf,'a') as log:
            log.write(str(cdtime)+" реквесты отправленны рисователям"+'\n')
    except:
        await bot.get_user(bot_master_id).send('не удалось отправить реквесты')
        with open(lf,'a') as log:
            log.write(str(cdtime)+" не удалось отправить реквесты рисователям"+'\n')
    return

#send direct message to author
@bot.command()
async def send_a(ctx):
    await ctx.author.send('hellow')
#send direct message to another
@bot.command()
async def send_to(ctx, member: discord.Member):
    await member.send(f'{member.name}, привет от '+ctx.author.name)

#получение данных в директе от участника (реакция на сообщения пользователя)
@bot.event
async def on_message(message):
    await bot.process_commands(message)
    msg = message.content.lower()
    bot_master = bot.get_user(bot_master_id)
    #приветсвие
    if  not message.guild and msg in hello_words:
        await message.channel.send(f'Привет, котик, я помощник секретного санты, принимаю реквесты. Проси помощь и я расскажу что делать')
        await send_report(message, ' здоровается с ботом')
    #ответ на ненужный вопрос
    if  not message.guild and msg in answer_words:
        await message.channel.send('42, человечек')
        await send_report(message, ' задаёт глупый вопрос')
    #Правила
    if  not message.guild and msg in rules_words:
        await message.channel.send(' О чём это:\n1. Участники тайно сообщают "Санте" (то есть мне, боту Тайного Санты) идею картинки, которую они хотят получить на Новый Год от неизвестного рисователя.\n2. В начале декабря идеи картинок тайным и случайным образом распределяются между участниками.\n3. Каждый участник тайно рисует картинку по желанию от другого неизвестного участника.\n4. Участники сдают выполненные картинки Санте (то есть мне, боту Тайного санты) через личные сообщения.\n5. Под Новый год выполненные картинки отправляются через бота Тайного Санты их желателям.\nПравила:\nТолько Санта знает, какую картинку загадал желающий.\nТолько Санта раздаёт (случайным образом) идеи картинок рисователям.\nСанта НЕ сообщает от кого поступило это желание.\nУчастники никому не сообщают, какую картинку они рисуют.\nДаже если рисователь догадался чьё это желание — ему не стоит сообщать об этом никому (особенно тому, чьё это желание).\nЕсли рисователю нужна дополнительная информация по реквесту - он может спросить @bespilotnik, чтобы тот связался с желающим для уточнения.\nКачество и уровень исполнения никак не нормируется, но некая законченность приветствуется.\nЖелательно следовать реквесту, потому что получить «Два треугольника красный и синий» по запросу «Cырно в камуфляже в лесу собирает Хурму», может быть немного досадно.\nПожалуйста, воздержитесь от совсем уж lewd реквестов в этом году. Снова.\nЖелательно предоставить референсы, если вы хотите увидеть что-то конкретное, иначе рисующий может неправильно вас понять. Референсы можно приложить к сообщению с реквестом картинкой или добавить позже, дополнив реквест, так же можно просто приложить ссылку\nВ первых числах декабря пожелания будут случайно распределены между участниками, которые незамедлительно приступят к рисованию.\nДедлайн сдачи рисунков 30 декабря (пожалуйста, не затягивайте до 23:59 31.12.22).\nВыдача подарков — 31 декабря.\n Есть вопросы? - Спросите у @bespilotnik')
    #herp
    if  not message.guild and msg in help_words:
        emb = discord.Embed(title = 'что говорить?')
        emb.add_field(name = 'Участвовать', value = 'чтобы записаться и сказать свой реквест')
        emb.add_field(name = 'Дополнить', value = 'если уже участвуешь, можно дополнить свой реквест')
        emb.add_field(name = 'Результат', value = 'если уже всё нарисовано чтобы отправить свой рисунок и пожелания')
        emb.add_field(name = 'Херп', value = 'помогай')
        emb.add_field(name = 'Правила', value = 'расскажу правила Секретного Санты')
        emb.add_field(name = '. _.', value = 'Можно использовать синонимы команд, но за их работу сложно поручиться. Реквест можно отправить текстом в сообщении, но не файлом. В сообщение к реквесту можно прикладывать картинки')
        await message.channel.send (embed = emb)
#Приём результатов от исполнителей
    if not message.guild and msg in recive_words :
        channel = message.channel
        requestor = message.author
        def check(message):
            return requestor == message.author
        if await check_participy(message, 'exequtor') != False:
            print ('уже участвует реквесты розданы')
            await channel.send('Скидывай свой рисунок(ки), можешь дополнить текстовым сообщением, которое получит желатель.\nЕсли у тебя заботливо пподготовлен архив то, пожалуйста, скинь его ссылкой')
            answer = await bot.wait_for('message', check=check)
            print('поиск строки исполнителя по исполнителю' + str(await check_participy(message, 'exequtor')) )
            await recive_result(answer, await check_participy(message, 'exequtor'))
            await message.channel.send('Результат принят, Секретный Санта')
        elif await check_participy(message, 'requestor') == False:
            print ('не участвует')
            await channel.send('Но ты ещё не участвуешь, хочешь участвовать?')
            answer = await bot.wait_for('message', check=check)
            if answer.content in yes_words:
                print('да получено')
                await channel.send('Диктуй свой реквест')
                req = await bot.wait_for('message', check=check)
                await add_req(message, req)
                await message.channel.send('Реквест принят, Секретный Санта')
            else:
                await channel.send('Тогда ладно')
                print('да пропущено')
            return
        else:
            print ('уже участвует реквесты НЕ розданы')
            await channel.send('Реквесты ещё не розданы, случайные картинки не принимаются')
            
#служебный подсчёт участников
    if  bot_master == message.author and not message.guild and msg == 'сколько участников':
        await message.channel.send('Уже участвует '+str(await number_of_participant()) + ' котов')
#служебное перемешивание участников и реквестов
    if  bot_master == message.author and not message.guild and msg == 'перемешай':
        await mixing_participant()
        await message.channel.send('Перемешиваю котов')
#служебная отправка реквестов исполнителям
    if  bot_master == message.author and not message.guild and msg == 'отправь реквесты':
        await send_request()
        await message.channel.send('Отправляю реквесты исполнителям')
#служебная отправка результатов желателям
    if  bot_master == message.author and not message.guild and msg == 'отправь результаты':
        await send_result()
        await message.channel.send('Отправляю результаты желателям')
#Запись в участники, проверка на участие и дополнение реквеста
    if  not message.guild and msg in paricipant_word:
        channel = message.channel
        requestor = message.author
        def check(message):
            return requestor == message.author
        if await check_participy(message,'requestor') != False:
            print ('уже участвует')
            await channel.send('Ты уже участвуешь, Хочешь что-то добавить?')
            answer = await bot.wait_for('message', check=check)
            if answer.content.lower() in yes_words:
                print('да получено')
                await channel.send('Диктуй своё дополнение')
                expreq = await bot.wait_for('message', check=check) 
                await expand_req(expreq, await check_participy(message, 'requestor'))
                await message.channel.send('Дополнение принято')
            else:
                await channel.send('Тогда ладно')
                print('да пропущено')
            return
        else: 
            print ('не участвует')
        await channel.send('Диктуй свой реквест')
        req = await bot.wait_for('message', check=check)
        await add_req(message, req)
        await message.channel.send('Реквест принят, Секретный Санта')
#Прямое дополнение реквеста с проверкой на участие (надо сделать отдельную функцию принятия реквеста и отдельную дополнения)
    if  not message.guild and msg in adding_words:
        channel = message.channel
        requestor = message.author
        def check(message):
            return requestor == message.author        
        if await check_participy(message,'requestor') == False:
            print ('ещё не участвует')
            await channel.send('Ты ещё не участвуешь, хочешь участвовать?')
            answer = await bot.wait_for('message', check=check)
            if answer.content in yes_words:
                print('да получено')
                await channel.send('Диктуй свой реквест')
                req = await bot.wait_for('message', check=check)
                await add_req(message, req)
                await message.channel.send('Реквест принят, Секретный Санта')
            else:
                await channel.send('Тогда ладно')
                print('да пропущено')
            return
        else: 
            print ('участвует')
            await channel.send('Диктуй своё дополнение')
            expreq = await bot.wait_for('message', check=check) 
            await expand_req(expreq, await check_participy(message,'requestor'))
            await message.channel.send('Дополнение принято')
#connect
with open(str(os.path.dirname(__file__))+'/pybot token.txt') as file: #Чтение токена из файла
    token = file.readline()
bot.run(token)